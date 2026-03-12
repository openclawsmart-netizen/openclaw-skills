#!/usr/bin/env python3
"""Generate Excel report from unified trade logs.

Source:
- data/trade_logs.json (JSON array)

Features:
- Flatten nested fields (including prior_trade_status)
- Keep P/L color highlighting
- Keep long-text wrapping
- Auto-send generated Excel to Telegram (graceful skip if env missing)
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, PatternFill


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT_DIR, "data", "trade_logs.json")
REPORT_DIR = os.path.join(ROOT_DIR, "reports")
TELEGRAM_SCRIPT = os.path.join(ROOT_DIR, "proactive-agent", "send_telegram.py")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _to_num(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _calc_actual_outcome(row: pd.Series) -> str:
    entry = _to_num(row.get("entry_price"))
    tp = _to_num(row.get("take_profit_price"))
    current = _to_num(row.get("prior_trade_status_current_price"))
    if current is None:
        current = _to_num(row.get("prior_current_price"))

    if entry is None or tp is None or current is None:
        return "NA"

    if tp > entry:
        return "WIN" if current >= entry else "LOSS"
    if tp < entry:
        return "WIN" if current <= entry else "LOSS"
    return "NA"


def _send_report_via_telegram(xlsx_path: str) -> int:
    if not os.path.exists(TELEGRAM_SCRIPT):
        print(f"[warn] Telegram script not found: {TELEGRAM_SCRIPT}", file=sys.stderr)
        return 0

    cmd = [
        sys.executable,
        TELEGRAM_SCRIPT,
        "--file",
        xlsx_path,
        "--caption",
        f"[report-generator] 交易報表：{os.path.basename(xlsx_path)}",
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)

    stdout = (proc.stdout or "").strip()
    stderr = (proc.stderr or "").strip()
    if stdout:
        print(stdout)
    if stderr:
        print(stderr, file=sys.stderr)

    if proc.returncode == 0:
        print("[ok] Auto-sent report via Telegram")
    elif proc.returncode == 2:
        print("[skip] Telegram env not configured; report generated locally")
        return 0
    else:
        print(f"[warn] Telegram send failed with code {proc.returncode}; report kept locally")

    return 0


def main() -> int:
    if not os.path.exists(JSON_PATH):
        print(f"[warn] JSON file not found: {JSON_PATH}")
        return 0

    try:
        with open(JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as exc:
        print(f"[error] Failed to read JSON: {exc}")
        return 1

    if not isinstance(data, list) or not data:
        print(f"[warn] No valid records in: {JSON_PATH}")
        return 0

    rows = [x for x in data if isinstance(x, dict)]
    if not rows:
        print(f"[warn] No valid object rows in: {JSON_PATH}")
        return 0

    df = pd.json_normalize(rows, sep="_")

    # 展開 prior_trade_status 巢狀欄位並友善命名
    rename_map = {
        "prior_trade_status_date": "prior_date",
        "prior_trade_status_entry_price": "prior_entry_price",
        "prior_trade_status_status": "prior_status",
        "prior_trade_status_current_price": "prior_current_price",
        "prior_trade_status_profit_loss_points": "prior_profit_loss_points",
    }
    exists = {k: v for k, v in rename_map.items() if k in df.columns}
    if exists:
        df = df.rename(columns=exists)

    df["actual_outcome"] = df.apply(_calc_actual_outcome, axis=1)

    os.makedirs(REPORT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(REPORT_DIR, f"trade_report_{ts}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="trade_logs")
        ws = writer.book["trade_logs"]

        header = [cell.value for cell in ws[1]]
        col_idx = {name: idx + 1 for idx, name in enumerate(header) if isinstance(name, str)}

        for text_col in ("analysis_reasoning", "error_review", "optimization_suggestion"):
            idx = col_idx.get(text_col)
            if not idx:
                continue
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).alignment = Alignment(wrap_text=True, vertical="top")

        for pnl_col in ("expected_profit_points", "prior_profit_loss_points", "prior_trade_status_profit_loss_points"):
            idx = col_idx.get(pnl_col)
            if not idx:
                continue
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=idx)
                val = _to_num(cell.value)
                if val is None:
                    continue
                if val > 0:
                    cell.fill = GREEN_FILL
                elif val < 0:
                    cell.fill = RED_FILL

    print(out_path)
    return _send_report_via_telegram(out_path)


if __name__ == "__main__":
    raise SystemExit(main())
